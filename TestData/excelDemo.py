import openpyxl

#Read
book = openpyxl.load_workbook("C:\\Users\\Lenovo\\PycharmProjects\\PythonDemo.xlsx")
sheet = book.active
Dict = {}
cellB1 = sheet.cell(row=1,column=2)
print(cellB1.value)

#Write
cellB2 = sheet.cell(row=2,column=2)
cellB2.value = "Cynthia"
print(cellB2.value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet["A5"].value)

#read the whole sheet
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i,column=j).value)

#read a specific row 2
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "Testcase2" :
        for j in range(1,sheet.max_column+1):
            print(sheet.cell(row=i,column=j).value)

#loading excel file to Dictionary
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "Testcase2" :
        for j in range(2,sheet.max_column+1):
            print(sheet.cell(row=i,column=j).value)
            Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value

print(Dict)
