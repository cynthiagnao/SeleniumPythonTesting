import openpyxl


class HomePageData:

    #a dictionnary
    test_HomePage_data = [{"firstname":"Cynthia","email":"cynthiagnao@gmail.com","gender":"Female"}, {"firstname":"Isaac","email":"mig@gmail.com","gender":"Male"}]

    @staticmethod
    def getTestData(test_case_name): #when a method is declared as static, self is no more required
        book = openpyxl.load_workbook("C:\\Users\\Lenovo\\PycharmProjects\\PythonDemo.xlsx")
        sheet = book.active
        Dict={}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    print(sheet.cell(row=i, column=j).value)
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]
